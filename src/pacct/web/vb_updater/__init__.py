"""
VB Updater: cruza descricoes de Virtual Bits entre o GLE do RDB (comment da
porta de saida do SYMBOL VBxxx) e o SCD (atributo `desc` do <ExtRef intAddr="VBxxx">
sob o IED correspondente).

Fluxo:
  1. Usuario faz upload de um RDB e de um SCD.
  2. App faz cross-match RDB <-> SCD usando `pacct.matchers.relay_scd`.
  3. Pra cada par casado, mostra um seletor de GLE + botao "Verify GLE comments".
  4. Clicar no botao abre uma pagina dedicada com a tabela de comparacao
     (VBxxx | comment do GLE | desc do SCD); celulas vazias mostram
     "<Without description>".

    templates/  landing.html e compare.html
"""

from __future__ import annotations

import json
import logging
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from html import escape
from pathlib import Path
from urllib.parse import parse_qs, quote, unquote, urlparse

from pacct.matchers import relay_scd as matcher
from pacct.parsers import rdb as rdb_loader
from pacct.parsers import scd as scd_loader
from pacct.parsers.gle import parse_gle
from pacct.parsers.rdb import RdbInfo
from pacct.paths import VB_UPDATER_TEMPLATES_DIR, is_within
from pacct.web import rdb_write
from pacct.web.project_files import library as filelib
from pacct.web.rdb_write import (
    resolve_gle_stream_path,
)
from pacct.web.rdb_write import (
    with_suffix_before_ext as _with_suffix_before_ext,
)
from pacct.web.session import SessionHandler
from pacct.web.xlsx_names import sanitize_sheet_name as _sanitize_sheet_name

_logger = logging.getLogger(__name__)


def load_template(name: str) -> str:
    """Read one template. Read at import time, like the GLV and the DNP map."""
    return (VB_UPDATER_TEMPLATES_DIR / name).read_text(encoding="utf-8")


# Limite generoso pra upload (usado so pelo /import-descriptions, que recebe
# o xlsx de volta -- o RDB e o SCD em si vem da biblioteca do projeto).
_SCD_MAX_BYTES = 200 * 1024 * 1024

# So consideramos VBs numericos (GOOSE virtual bits). Math variables tipo
# VBY, VBZ, VBRMS, etc. ficam de fora.
_VB_NUMERIC_RE = re.compile(r"^VB(\d+)$", re.IGNORECASE)

# Diretorio onde os SCDs uploadados sao salvos. Fica em cache/ pra ser
# gitignored automaticamente.
# Uploads e saidas vivem em cache/sessions/<sid>/ (ver pacct.web.session).

_EMPTY_LABEL = "&lt;Without description&gt;"

# Comment usado pra renomear VBs que aparecem no SCD com `desc` vazio
# (i.e., declarados como ExtRef mas sem descricao). Convencao: "reserva"
# (spare) -- indicam um VB previsto mas nao usado.
_RESERVA_LABEL = "reserva"


# -----------------------------------------------------------------------------
# Extractors
# -----------------------------------------------------------------------------

@dataclass(frozen=True)
class GleVbInstance:
    """Uma ocorrencia de um SYMBOL VBnnn no GLE."""
    page: str          # nome da <page> que contem o elemento
    element_id: str    # id do <element type="SYMBOL"> que envolve o logic_element
    comment: str       # comment da primeira <port> nao vazia (string vazia se nenhuma)


def extract_vb_instances_from_gle(gle_path: Path) -> dict[str, list[GleVbInstance]]:
    """Le um GLE.xml e retorna {VBnnn: [GleVbInstance, ...]}.

    Um mesmo VB pode aparecer varias vezes no diagrama (em paginas diferentes
    ou ate dentro da mesma pagina); cada ocorrencia gera uma instancia.

    So pega SYMBOLs cujo `physical_instance_name` casa com `VB\\d+`. O
    comment de cada instancia eh lido da primeira <port> nao vazia. Se nenhuma
    porta tiver comment, a instancia ainda eh registrada com `comment=""`.
    """
    out: dict[str, list[GleVbInstance]] = {}
    try:
        # `parse_gle` ja lida com encoding misto (utf-8 declarado mas
        # conteudo latin-1) -- bug comum em GLEs exportados pelo QuickSet.
        root = parse_gle(gle_path)
    except (OSError, ET.ParseError, UnicodeDecodeError) as e:
        _logger.warning("erro lendo GLE %s: %s", gle_path, e)
        return out

    # Iteramos por pagina pra preservar a localizacao de cada SYMBOL.
    for page in root.iter("page"):
        page_name = (page.attrib.get("name") or "").strip()
        for el in page.iter("element"):
            if el.attrib.get("type") != "SYMBOL":
                continue
            elem_id = (el.attrib.get("id") or "").strip()
            for logic_el in el.iter("logic_element"):
                if logic_el.attrib.get("type") != "SYMBOL":
                    continue
                name = (logic_el.attrib.get("physical_instance_name") or "").strip()
                m = _VB_NUMERIC_RE.match(name)
                if not m:
                    continue
                key = f"VB{int(m.group(1))}"
                # Acha a primeira <port> com comment nao vazio (qualquer index).
                # GLEs podem ter dois blocos <ports> seguidos -- o segundo eh
                # o que carrega os dados.
                port_comment = ""
                for port in logic_el.iter("port"):
                    comment_el = port.find("comment")
                    if comment_el is not None and (comment_el.text or "").strip():
                        port_comment = (comment_el.text or "").strip()
                        break
                out.setdefault(key, []).append(GleVbInstance(
                    page=page_name, element_id=elem_id, comment=port_comment,
                ))
    return out


def extract_vb_descriptions_from_scd_ied(scd_path: Path, ied_name: str) -> dict[str, str]:
    """Le um SCD e retorna {VBxxx: desc_do_ExtRef} para o IED informado.

    Procura todos os <ExtRef> que estao dentro de <IED name=ied_name> e cujo
    `intAddr` casa com VBnnn. Mapeia VBnnn -> primeiro `desc` nao vazio
    encontrado (varios ExtRefs podem referenciar o mesmo intAddr).
    """
    out: dict[str, str] = {}
    try:
        tree = ET.parse(str(scd_path))
    except (OSError, ET.ParseError) as e:
        _logger.warning("erro lendo SCD %s: %s", scd_path, e)
        return out

    root = tree.getroot()
    target_ied = None
    # _iter_local ignora namespace, entao 'IED' casa com '{ns}IED'.
    for el in scd_loader._iter_local(root, "IED"):
        if el.attrib.get("name") == ied_name:
            target_ied = el
            break
    if target_ied is None:
        _logger.info("IED %r nao encontrado no SCD %s", ied_name, scd_path)
        return out

    for ext in scd_loader._iter_local(target_ied, "ExtRef"):
        addr = (ext.attrib.get("intAddr") or "").strip()
        m = _VB_NUMERIC_RE.match(addr)
        if not m:
            continue
        key = f"VB{int(m.group(1))}"
        desc = (ext.attrib.get("desc") or "").strip()
        existing = out.get(key, "")
        if existing and not desc:
            continue
        if desc or key not in out:
            out[key] = desc
    return out


# Campos do ExtRef que descrevem a "assinatura" GOOSE. Ordem usada na composicao
# do Signal: <iedName>/<srcLDInst>/<srcLNClass>/<srcCBName>.<ldInst>.<prefix><lnClass><lnInst>.<doName>.<daName>
_EXTREF_FIELDS = (
    "iedName", "srcLDInst", "srcLNClass", "srcCBName",
    "ldInst", "prefix", "lnClass", "lnInst", "doName", "daName",
)


def _format_extref_signal(attrs: dict[str, str]) -> str:
    """Monta a string de Signal a partir dos atributos de um ExtRef.

    Retorna "" se nao houver `iedName` -- esses ExtRefs sao placeholders
    (intAddr definido mas sem subscription real).
    """
    ied = (attrs.get("iedName") or "").strip()
    if not ied:
        return ""
    src_ld = (attrs.get("srcLDInst") or "").strip()
    src_ln = (attrs.get("srcLNClass") or "").strip()
    src_cb = (attrs.get("srcCBName") or "").strip()
    ld = (attrs.get("ldInst") or "").strip()
    prefix = (attrs.get("prefix") or "").strip()
    ln_cls = (attrs.get("lnClass") or "").strip()
    ln_inst = (attrs.get("lnInst") or "").strip()
    do = (attrs.get("doName") or "").strip()
    da = (attrs.get("daName") or "").strip()
    # `prefix` eh opcional no SCL -- so concatena com lnClass+lnInst se existir.
    ln_token = f"{prefix}{ln_cls}{ln_inst}"
    return f"{ied}/{src_ld}/{src_ln}/{src_cb}.{ld}.{ln_token}.{do}.{da}"


def extract_vb_extref_rows_from_scd_ied(
    scd_path: Path, ied_name: str,
) -> list[dict]:
    """Le um SCD e retorna uma lista de dicts {vb, signal, desc} -- uma linha
    por VBxxx encontrado nos ExtRef do IED informado.

    Se um mesmo VB aparece em varios ExtRefs (tipico: 1 placeholder + 1 com
    subscription), preferimos o que tem `iedName` populado (signal real). Se
    nenhum tem, mantemos o primeiro (signal vazio).
    """
    rows_by_vb: dict[str, dict] = {}
    try:
        tree = ET.parse(str(scd_path))
    except (OSError, ET.ParseError) as e:
        _logger.warning("erro lendo SCD %s: %s", scd_path, e)
        return []

    root = tree.getroot()
    target_ied = None
    for el in scd_loader._iter_local(root, "IED"):
        if el.attrib.get("name") == ied_name:
            target_ied = el
            break
    if target_ied is None:
        return []

    for ext in scd_loader._iter_local(target_ied, "ExtRef"):
        addr = (ext.attrib.get("intAddr") or "").strip()
        m = _VB_NUMERIC_RE.match(addr)
        if not m:
            continue
        vb = f"VB{int(m.group(1))}"
        attrs = {k: ext.attrib.get(k, "") for k in _EXTREF_FIELDS}
        signal = _format_extref_signal(attrs)
        desc = (ext.attrib.get("desc") or "").strip()
        row = {"vb": vb, "signal": signal, "desc": desc}
        existing = rows_by_vb.get(vb)
        if existing is None:
            rows_by_vb[vb] = row
            continue
        # Prefere a linha com signal preenchido (subscription real).
        if not existing["signal"] and signal:
            rows_by_vb[vb] = row
    return sorted(rows_by_vb.values(), key=lambda r: int(r["vb"][2:]))


# -----------------------------------------------------------------------------
# Writers: GLE (dentro do RDB) e SCD
# -----------------------------------------------------------------------------

# Casa um <logic_element type="SYMBOL" ... physical_instance_name="VBnnn" ...>
# ... </logic_element> em bytes. group(1) = atributos (pra pegar nome), group(2)
# = corpo (entre as tags).
_GLE_VB_BLOCK_RE = re.compile(
    rb'(<logic_element\s+type="SYMBOL"[^>]*physical_instance_name="VB(\d+)"[^>]*>)'
    rb'(.*?)'
    rb'(</logic_element>)',
    re.DOTALL,
)

# Casa <port ...> ... <comment>TEXT</comment> ... </port> dentro de um bloco
# de logic_element. Limita-se a 1 substituicao por bloco (a primeira porta).
_GLE_PORT_COMMENT_RE = re.compile(
    rb'(<port\b[^>]*>\s*<comment>)([^<]*)(</comment>\s*</port>)',
    re.DOTALL,
)

# Casa <ExtRef ... intAddr="VBnnn" ... /> dentro de um IED. group(1) = atributos
# antes de intAddr, group(2) = numero do VB, group(3) = atributos depois.
# IMPORTANTE: o character class precisa permitir `/` -- valores de `desc` no
# campo costumam ter '/' (ex.: "50/62BF LT1 UPC1"); excluir `/` aqui faria a
# regex ignorar silenciosamente esses ExtRefs.
_SCD_EXTREF_RE = re.compile(
    rb'<ExtRef\b(?P<attrs>[^>]*?intAddr="VB(?P<num>\d+)"[^>]*?)(?P<close>/?>)',
    re.DOTALL,
)

_SCD_DESC_ATTR_RE = re.compile(rb'desc="[^"]*"')
_SCD_INTADDR_ATTR_RE = re.compile(rb'intAddr="VB\d+"')


def _xml_attr_escape(s: str) -> str:
    """Escapa caracteres especiais para uso dentro de um atributo XML "..."."""
    return (s.replace("&", "&amp;")
             .replace("<", "&lt;")
             .replace('"', "&quot;"))


def _substitute_vb_comments_in_gle_bytes(
    raw: bytes, new_comments: dict[str, str],
) -> tuple[bytes, dict[str, int]]:
    """Atualiza o comment da PRIMEIRA porta de cada SYMBOL VBnnn no GLE.

    Retorna (novos_bytes, stats). `stats` tem chaves:
      - "updated":   numero de instancias VBnnn substituidas
      - "skipped":   VBnnn em `new_comments` que nao tiveram porta-comment encontrada
      - "untouched": VBnnn no GLE mas sem entry em `new_comments`
    """
    stats = {"updated": 0, "skipped": 0, "untouched": 0}

    def replace_block(m: re.Match) -> bytes:
        head = m.group(1)
        num_str = m.group(2).decode("ascii")
        body = m.group(3)
        tail = m.group(4)
        key = f"VB{int(num_str)}"
        new_text = new_comments.get(key)
        if new_text is None:
            stats["untouched"] += 1
            return m.group(0)
        # Encoda em latin-1 pra casar com o restante do GLE (Quickset escreve
        # latin-1 embora declare utf-8 no header).
        new_text_bytes = new_text.encode("latin-1", errors="replace")
        new_body, count = _GLE_PORT_COMMENT_RE.subn(
            lambda pm: pm.group(1) + new_text_bytes + pm.group(3),
            body,
            count=1,
        )
        if count == 0:
            stats["skipped"] += 1
            return m.group(0)
        stats["updated"] += 1
        return head + new_body + tail

    new_raw = _GLE_VB_BLOCK_RE.sub(replace_block, raw)
    return new_raw, stats


def _update_scd_extrefs_for_ied(
    raw: bytes, ied_name: str, new_descs: dict[str, str],
) -> tuple[bytes, dict[str, int]]:
    """Atualiza o atributo `desc` de cada <ExtRef intAddr="VBnnn"> dentro do
    IED `ied_name`. Se o ExtRef nao tem `desc=`, insere logo apos `intAddr=`.

    Retorna (novos_bytes, stats). Bytes podem mudar de tamanho (SCD eh XML
    puro, sem constraint).
    """
    stats = {"updated": 0, "inserted": 0, "untouched": 0}

    # Acha o bloco do IED especifico.
    ied_re = re.compile(
        rb'<IED\b[^>]*name="' + re.escape(ied_name.encode("utf-8")) + rb'"[^>]*>',
    )
    open_match = ied_re.search(raw)
    if not open_match:
        return raw, stats
    # O </IED> correspondente: procuramos do open_match em diante. Como o
    # SCD nao tem IEDs aninhados, basta o proximo `</IED>`.
    close_idx = raw.find(b"</IED>", open_match.end())
    if close_idx == -1:
        return raw, stats
    inner_start = open_match.end()
    inner_end = close_idx

    def replace_extref(em: re.Match) -> bytes:
        full = em.group(0)
        num_str = em.group("num").decode("ascii")
        key = f"VB{int(num_str)}"
        new = new_descs.get(key)
        if not new:
            stats["untouched"] += 1
            return full
        new_attr = ('desc="' + _xml_attr_escape(new) + '"').encode("utf-8")
        if _SCD_DESC_ATTR_RE.search(full):
            new_full, n = _SCD_DESC_ATTR_RE.subn(new_attr, full, count=1)
            if n > 0:
                stats["updated"] += 1
                return new_full
        # Insere apos intAddr="..."
        def insert(im: re.Match) -> bytes:
            return im.group(0) + b" " + new_attr
        new_full, n = _SCD_INTADDR_ATTR_RE.subn(insert, full, count=1)
        if n > 0:
            stats["inserted"] += 1
            return new_full
        return full  # caso degenerado

    inner_new = _SCD_EXTREF_RE.sub(replace_extref, raw[inner_start:inner_end])
    return raw[:inner_start] + inner_new + raw[inner_end:], stats


# -----------------------------------------------------------------------------
# Orquestradores: copia arquivo + aplica substituicoes + grava `_comments_updated`
# -----------------------------------------------------------------------------

def _gle_stream_path(extract_dir: Path, relay_name: str, gle_name: str,
                     gle_fs_path: Path) -> list[str]:
    """O caminho do stream do GLE dentro do OLE, com o palpite otimista.

    Quando o arquivo extraido nao esta sob `extract_dir` -- o que nao acontece
    com um GLE vindo do `RdbInfo` -- monta o caminho a partir dos nomes, que e'
    o layout que todo RDB do corpus usa.
    """
    return resolve_gle_stream_path(
        extract_dir, gle_fs_path,
        fallback=["Relays", relay_name, "Misc", f"{gle_name}.gle"],
    )


def _new_comments_from_scd(scd_path: Path, ied_name: str) -> tuple[dict, int, int]:
    """As descricoes que o SCD manda pro GLE, mais as duas contagens do relatorio.

    VB que existe como ExtRef mas com `desc` vazia vira "reserva": e' um slot
    previsto e nao usado, e deixar o comentario antigo la seria pior -- ele
    descreveria um sinal que nao existe mais.
    """
    new_comments: dict[str, str] = {}
    reserva = 0
    with_desc = 0
    for vb, desc in extract_vb_descriptions_from_scd_ied(scd_path, ied_name).items():
        if desc:
            new_comments[vb] = desc
            with_desc += 1
        else:
            new_comments[vb] = _RESERVA_LABEL
            reserva += 1
    return new_comments, with_desc, reserva


def _apply_stats(sub_stats: dict, with_desc: int, reserva: int,
                 original_size: int) -> dict:
    return {
        "instances_updated": sub_stats["updated"],
        "vbs_not_found_in_gle": sub_stats["skipped"],
        "vbs_in_gle_not_in_scd": sub_stats["untouched"],
        "vbs_in_scd_with_desc": with_desc,
        "vbs_in_scd_renamed_to_reserva": reserva,
        "reserva_label": _RESERVA_LABEL,
        "original_stream_bytes": original_size,
    }


def update_rdb_with_scd_descs(
    *,
    rdb_path: Path,
    extract_dir: Path,
    relay_name: str,
    gle_name: str,
    gle_fs_path: Path,
    scd_path: Path,
    ied_name: str,
    output_path: Path,
    job=None,
) -> dict:
    """Gera `output_path` (copia do RDB) com o GLE selecionado atualizado:
    comments das portas VBnnn substituidos pelos `desc` do SCD.

    Uma passada de leitura, depois uma gravacao. `rdb_write.write_streams`
    escolhe entre reescrever o stream no lugar (quando o tamanho bate) e
    reconstruir o container (quando nao bate), e grava atomicamente. Antes o
    stream tinha que caber no tamanho original e o excedente era comprado
    colapsando whitespace no `.gle` inteiro -- ver `pacct/web/rdb_write.py`.

    Com `ok: False`, nenhum arquivo foi escrito.
    """
    new_comments, with_desc, reserva = _new_comments_from_scd(scd_path, ied_name)
    if not new_comments:
        return {
            "ok": False,
            "error": f"SCD nao tem nenhum ExtRef VBnnn para o IED {ied_name!r}.",
        }

    original = Path(gle_fs_path).read_bytes()
    updated, sub_stats = _substitute_vb_comments_in_gle_bytes(original, new_comments)
    stream_parts = _gle_stream_path(extract_dir, relay_name, gle_name, gle_fs_path)

    try:
        method = rdb_write.write_streams(
            rdb_path, output_path, {tuple(stream_parts): updated}, job=job)
    except rdb_write.RdbWriteError as e:
        return {"ok": False, "error": str(e),
                "stats": _apply_stats(sub_stats, with_desc, reserva, len(original))}

    return {
        "ok": True,
        "output_path": str(output_path),
        "method": method,
        "stream": "/".join(stream_parts),
        "stats": _apply_stats(sub_stats, with_desc, reserva, len(original)),
    }


def update_rdb_with_scd_descs_batch(
    *,
    rdb_path: Path,
    extract_dir: Path,
    scd_path: Path,
    selections: list[dict],
    output_path: Path,
    job=None,
) -> dict:
    """Aplica `update_rdb_with_scd_descs` para varios reles num unico RDB.

    `selections` eh uma lista de dicts {relay, ied, gle_name, gle_fs_path}.

    Tudo-ou-nada, e essa e' a mudanca que importa aqui. Antes cada selecao era
    gravada no RDB de saida assim que ficava pronta: se a terceira falhava, as
    duas primeiras ja estavam no arquivo, a resposta ainda dizia `ok: True` e
    esse RDB meio-aplicado entrava no acervo do projeto do mesmo jeito que um
    inteiro. Depois que sai daqui nao da mais pra saber a diferenca.

    Retorna {ok, output_path, method, succeeded, failed, results}. Com
    `ok: False`, NENHUM arquivo foi escrito.
    """
    results: list[dict] = []
    streams: dict[tuple[str, ...], bytes] = {}

    if job:
        job.stage("Conferindo as seleções", 10)
    for sel in selections:
        relay = sel["relay"]
        ied = sel["ied"]
        gle_name = sel["gle_name"]
        gle_fs_path = Path(sel["gle_fs_path"])
        entry_result: dict = {"relay": relay, "ied": ied, "gle": gle_name}

        new_comments, with_desc, reserva = _new_comments_from_scd(scd_path, ied)
        if not new_comments:
            entry_result["ok"] = False
            entry_result["error"] = f"SCD sem ExtRef VBnnn para IED {ied!r}."
            results.append(entry_result)
            continue

        if not gle_fs_path.is_file():
            entry_result["ok"] = False
            entry_result["error"] = f"GLE nao encontrado: {gle_fs_path.name}"
            results.append(entry_result)
            continue

        stream_parts = tuple(_gle_stream_path(extract_dir, relay, gle_name,
                                              gle_fs_path))
        original = gle_fs_path.read_bytes()
        updated, sub_stats = _substitute_vb_comments_in_gle_bytes(
            original, new_comments)
        streams[stream_parts] = updated
        entry_result["ok"] = True
        entry_result["stats"] = _apply_stats(sub_stats, with_desc, reserva,
                                             len(original))
        results.append(entry_result)

    succeeded = sum(1 for r in results if r.get("ok"))
    failed = len(results) - succeeded
    if failed or not streams:
        return {
            "ok": False,
            "error": (f"{failed} de {len(results)} seleção(ões) falharam; "
                      "nenhum RDB foi gravado."
                      if failed else "Nenhuma seleção produziu alteração."),
            "succeeded": succeeded,
            "failed": failed,
            "results": results,
        }

    try:
        method = rdb_write.write_streams(rdb_path, output_path, streams, job=job)
    except rdb_write.RdbWriteError as e:
        return {"ok": False, "error": str(e), "succeeded": succeeded,
                "failed": failed, "results": results}

    return {
        "ok": True,
        "output_path": str(output_path),
        "method": method,
        "succeeded": succeeded,
        "failed": failed,
        "results": results,
    }


def update_scd_with_gle_comments(
    *,
    scd_path: Path,
    ied_name: str,
    gle_path: Path,
    output_path: Path,
) -> dict:
    """Gera `output_path` (copia do SCD) com os <ExtRef desc=> do IED `ied_name`
    atualizados a partir dos comments das portas VBnnn do GLE.

    Quando o mesmo VB aparece varias vezes no GLE com comments diferentes,
    usa o PRIMEIRO comment nao vazio encontrado (e inclui um aviso).
    """
    gle_map = extract_vb_instances_from_gle(gle_path)

    new_descs: dict[str, str] = {}
    inconsistent: list[str] = []
    for vb, insts in gle_map.items():
        first_non_empty = ""
        all_comments: set[str] = set()
        for inst in insts:
            if inst.comment:
                if not first_non_empty:
                    first_non_empty = inst.comment
                all_comments.add(inst.comment)
        if first_non_empty:
            new_descs[vb] = first_non_empty
        if len(all_comments) > 1:
            inconsistent.append(vb)

    if not new_descs:
        return {"ok": False, "error": "GLE nao tem nenhum comment nao vazio para VBs."}

    scd_bytes = scd_path.read_bytes()
    updated, stats = _update_scd_extrefs_for_ied(scd_bytes, ied_name, new_descs)
    if stats["updated"] + stats["inserted"] == 0:
        return {
            "ok": False,
            "error": f"IED {ied_name!r} nao encontrado no SCD ou sem ExtRef VBnnn.",
        }

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(updated)

    return {
        "ok": True,
        "output_path": str(output_path),
        "stats": {
            "extrefs_updated": stats["updated"],
            "extrefs_inserted_desc": stats["inserted"],
            "vbs_unchanged_no_match": stats["untouched"],
            "vbs_in_gle_with_comment": len(new_descs),
            "vbs_with_inconsistent_gle_comments": inconsistent,
            "original_bytes": len(scd_bytes),
            "output_bytes": len(updated),
        },
    }


# -----------------------------------------------------------------------------
# Excel I/O: export descriptions to xlsx, parse edited xlsx back
# -----------------------------------------------------------------------------

# Sheet names no Excel: max 31 chars, sem `:\\/?*[]`. Sanitizamos pra evitar
# erros do openpyxl ao salvar.

# Header esperado nas linhas 1 e 3 do sheet. Marcador na A1 permite identificar
# que o arquivo veio do export (e nao um xlsx aleatorio).
_XLSX_IED_MARKER = "IED:"
_XLSX_HEADER_VB = "VB"
_XLSX_HEADER_SIGNAL = "Signal"
_XLSX_HEADER_DESC = "Description"


def build_vb_descriptions_xlsx(
    *, scd_path: Path, ied_names: list[str], rdb_by_ied: dict[str, str] | None = None,
) -> bytes:
    """Gera um .xlsx (bytes) com uma aba por IED em `ied_names`.

    Estrutura por aba:
      A1: "IED:"           B1: <ied_name>            (marcador + identificacao)
      A2: "Relay:"         B2: <rdb_relay_name>      (so se rdb_by_ied passado)
      Linha 3 em branco
      Linha 4: cabecalhos  ["VB", "Signal", "Description"]
      Linha 5+: dados

    O importer le o IED de B1 (nao do nome da aba), entao o usuario pode
    renomear/reordenar abas sem quebrar o re-import.
    """
    # Import local: openpyxl eh dependencia opcional usada so neste fluxo.
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # Remove a sheet default; vamos criar as nossas.
    default = wb.active
    wb.remove(default)

    used: set[str] = set()
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF1F6FEB")
    meta_font = Font(bold=True)

    for ied in ied_names:
        rows = extract_vb_extref_rows_from_scd_ied(scd_path, ied)
        ws = wb.create_sheet(title=_sanitize_sheet_name(ied, used))
        ws["A1"] = _XLSX_IED_MARKER
        ws["B1"] = ied
        ws["A1"].font = meta_font
        rdb_name = (rdb_by_ied or {}).get(ied)
        if rdb_name:
            ws["A2"] = "Relay:"
            ws["B2"] = rdb_name
            ws["A2"].font = meta_font
        # Header na linha 4.
        for col, val in enumerate(
            (_XLSX_HEADER_VB, _XLSX_HEADER_SIGNAL, _XLSX_HEADER_DESC), start=1,
        ):
            c = ws.cell(row=4, column=col, value=val)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="left")
        for i, row in enumerate(rows, start=5):
            ws.cell(row=i, column=1, value=row["vb"])
            ws.cell(row=i, column=2, value=row["signal"])
            ws.cell(row=i, column=3, value=row["desc"])
        # Auto-width grosseiro (openpyxl nao tem auto-fit nativo).
        widths = {1: 10, 2: 90, 3: 60}
        for col_idx, w in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = w
        # Freeze pane abaixo do header.
        ws.freeze_panes = "A5"

    if not wb.sheetnames:
        # openpyxl exige pelo menos uma sheet pra salvar.
        wb.create_sheet(title="empty")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_vb_descriptions_xlsx(xlsx_bytes: bytes) -> dict[str, dict[str, str]]:
    """Le um xlsx gerado por `build_vb_descriptions_xlsx` (ou editado) e retorna
    {ied_name: {VBxxx: nova_descricao}}.

    Regras:
      - IED de cada aba lido de B1 (nao do titulo). Abas sem marcador "IED:"
        em A1 ou B1 vazio sao ignoradas.
      - Apos a linha de header (4), linhas com VB vazio ou descricao vazia
        sao puladas (empty = nao mexer).
      - VBs sao normalizados pra VB{int(n)} (case-insensitive).
      - Se o mesmo VB aparece duas vezes na mesma aba, a ultima descricao vence.
    """
    from io import BytesIO

    from openpyxl import load_workbook

    wb = load_workbook(filename=BytesIO(xlsx_bytes), data_only=True, read_only=True)
    out: dict[str, dict[str, str]] = {}
    try:
        for ws in wb.worksheets:
            # Verifica marcador A1 e le IED de B1.
            a1 = ws.cell(row=1, column=1).value
            b1 = ws.cell(row=1, column=2).value
            if not isinstance(a1, str) or a1.strip() != _XLSX_IED_MARKER:
                continue
            ied = (str(b1).strip() if b1 is not None else "")
            if not ied:
                continue
            per_ied: dict[str, str] = out.setdefault(ied, {})
            # Itera a partir da linha 5 (data).
            for row in ws.iter_rows(min_row=5, max_col=3, values_only=True):
                if row is None:
                    continue
                vb_val, _signal, desc_val = (row + (None, None, None))[:3]
                if vb_val is None:
                    continue
                vb_str = str(vb_val).strip()
                m = _VB_NUMERIC_RE.match(vb_str)
                if not m:
                    continue
                key = f"VB{int(m.group(1))}"
                if desc_val is None:
                    continue
                desc = str(desc_val).strip()
                if not desc:
                    continue
                per_ied[key] = desc
    finally:
        wb.close()
    # Remove IEDs sem nenhuma desc valida.
    return {k: v for k, v in out.items() if v}


def update_scd_with_descriptions_multi(
    *, scd_path: Path, descriptions_by_ied: dict[str, dict[str, str]],
    output_path: Path,
) -> dict:
    """Aplica novas descricoes a varios IEDs em um unico SCD e grava em
    `output_path`. Reusa `_update_scd_extrefs_for_ied` por IED.

    Retorna dict com per_ied stats e totais.
    """
    raw = scd_path.read_bytes()
    per_ied_stats: dict[str, dict] = {}
    skipped_no_ied: list[str] = []
    totals = {"updated": 0, "inserted": 0, "untouched": 0}

    for ied, vb_map in descriptions_by_ied.items():
        if not vb_map:
            continue
        new_raw, stats = _update_scd_extrefs_for_ied(raw, ied, vb_map)
        if stats["updated"] + stats["inserted"] == 0 and stats["untouched"] == 0:
            # IED nao encontrado ou sem ExtRef VBnnn.
            skipped_no_ied.append(ied)
            continue
        raw = new_raw
        per_ied_stats[ied] = stats
        totals["updated"] += stats["updated"]
        totals["inserted"] += stats["inserted"]
        totals["untouched"] += stats["untouched"]

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(raw)
    return {
        "ok": True,
        "output_path": str(output_path),
        "per_ied": per_ied_stats,
        "skipped_ieds": skipped_no_ied,
        "totals": totals,
        "output_bytes": len(raw),
    }


# -----------------------------------------------------------------------------
# Estado da sessao (modulo-global, uma instancia por processo)
# -----------------------------------------------------------------------------

@dataclass
class _SessionState:
    rdb: RdbInfo | None = None
    scd_path: Path | None = None
    scd_name: str | None = None
    match_report: matcher.MatchReport | None = None


def _maybe_match(st: _SessionState) -> None:
    """Roda o cross-match se ja temos RDB e SCD. Chamar sob o lock da sessao."""
    if st.rdb is None or st.scd_path is None:
        st.match_report = None
        return
    extract_dir = st.rdb.extract_dir
    try:
        st.match_report = matcher.compare_relays_to_scd(
            st.rdb.relays, extract_dir, st.scd_path,
        )
    except Exception as e:
        _logger.exception("falha no cross-match: %s", e)
        st.match_report = None


def _state_payload(st: _SessionState) -> dict:
    d = {
        "has_rdb": st.rdb is not None,
        "has_scd": st.scd_path is not None,
        "rdb_name": st.rdb.display_name if st.rdb else None,
        "scd_name": st.scd_name,
        "matches": [],
        "unmatched_rdb": [],
        "unmatched_scd": [],
    }
    if st.match_report is not None:
        d["matches"] = [m.to_dict() for m in st.match_report.matched]
        d["unmatched_rdb"] = [u.to_dict() for u in st.match_report.unmatched_rdb]
        d["unmatched_scd"] = [u.to_dict() for u in st.match_report.unmatched_scd]
    # GLEs por relay (pra popular o <select> de cada linha)
    if st.rdb is not None:
        gles_by_relay: dict[str, list[str]] = {}
        for r in st.rdb.relays:
            gles_by_relay[r.name] = [g.name for g in r.gles]
        d["gles_by_relay"] = gles_by_relay
    else:
        d["gles_by_relay"] = {}
    return d


# -----------------------------------------------------------------------------
# Comparison page renderer
# -----------------------------------------------------------------------------

def _render_compare_page(rdb_relay: str, ied_name: str, gle_name: str,
                         gle_path: Path, scd_path: Path) -> str:
    """Gera o HTML completo da pagina de comparacao.

    Uma linha por instancia de VB no GLE. Se um VB aparece N vezes no GLE,
    sao geradas N linhas (todas comparadas contra a mesma desc do SCD). Se
    nao aparece no GLE mas existe no SCD, vira 1 linha com GLE vazio.
    """
    gle_map = extract_vb_instances_from_gle(gle_path)
    scd_map = extract_vb_descriptions_from_scd_ied(scd_path, ied_name)

    all_vbs = sorted(
        set(gle_map.keys()) | set(scd_map.keys()),
        key=lambda k: int(k[2:]),
    )

    rows = []
    equal_count = 0
    diff_count = 0

    def _row(vb: str, gle_comment: str, scd_desc: str, loc: str) -> str:
        is_diff = (gle_comment != scd_desc)
        nonlocal equal_count, diff_count
        if is_diff:
            diff_count += 1
        else:
            equal_count += 1
        g_html = escape(gle_comment) if gle_comment else _EMPTY_LABEL
        s_html = escape(scd_desc) if scd_desc else _EMPTY_LABEL
        g_cls = "cell" + ("" if gle_comment else " empty")
        s_cls = "cell" + ("" if scd_desc else " empty")
        loc_html = f'<div class="loc">{escape(loc)}</div>' if loc else ""
        cls = ' class="diff"' if is_diff else ""
        return (
            f'<tr{cls}>'
            f'<td class="vb">{escape(vb)}{loc_html}</td>'
            f'<td class="{g_cls}">{g_html}</td>'
            f'<td class="{s_cls}">{s_html}</td>'
            f'</tr>'
        )

    for vb in all_vbs:
        instances = gle_map.get(vb, [])
        scd_desc = scd_map.get(vb, "")
        if not instances:
            # VB existe no SCD mas nao no GLE.
            rows.append(_row(vb, "", scd_desc, "(ausente no GLE)"))
            continue
        for inst in instances:
            parts = [p for p in (inst.page, f"#{inst.element_id}" if inst.element_id else "") if p]
            loc = " / ".join(parts)
            rows.append(_row(vb, inst.comment, scd_desc, loc))

    total_rows = equal_count + diff_count
    summary = (
        f'<span class="ok">{equal_count} iguais</span> &nbsp; '
        f'<span class="warn">{diff_count} divergentes</span> &nbsp; '
        f'<span class="muted">{total_rows} instancia(s) &middot; '
        f'{len(all_vbs)} VB(s) unico(s)</span>'
    )

    body_rows = "\n".join(rows) if rows else (
        '<tr><td colspan="3" class="muted" style="text-align:center;padding:24px">'
        'Nenhum VB encontrado nos dois lados.</td></tr>'
    )

    return COMPARE_HTML_TEMPLATE.format(
        rdb_relay=escape(rdb_relay),
        ied_name=escape(ied_name),
        gle_name=escape(gle_name),
        summary=summary,
        rows=body_rows,
    )


# -----------------------------------------------------------------------------
# HTML
# -----------------------------------------------------------------------------

LANDING_HTML = load_template("landing.html")

# A navegacao numerada e' a mesma das nove telas -- mora em theme.py.


COMPARE_HTML_TEMPLATE = load_template("compare.html")



# -----------------------------------------------------------------------------
# Server
# -----------------------------------------------------------------------------

def build_vb_updater_handler(logger: logging.Logger, sessions) -> type:
    """Devolve a classe de handler do VB Updater.

    Nao sobe servidor: quem serve e' o dispatcher unico de `pacct.web.mount`,
    que monta esse handler em `/vb-updater/`. Estado e uploads ficam por sessao
    (`self.sess()` / `self.sdir()`), nao por processo.
    """

    class Handler(SessionHandler):
        session_key = "vb-updater"
        state_factory = _SessionState
        server_sessions = sessions

        def do_GET(self):
            parsed = urlparse(self.path)
            path = parsed.path
            qs = parse_qs(parsed.query)

            if path in ("/", "/index.html"):
                self._send(200, LANDING_HTML, "text/html; charset=utf-8")
                return
            if path == "/vb-state":
                # Sentinela usado por home/dashboard pra detectar que essa
                # ferramenta esta no ar.
                self._send_json(200, {"ok": True})
                return
            if path == "/state":
                self._send_json(200, _state_payload(self.sess()))
                return
            if path == "/download":
                # Serve arquivo gerado pelo /apply como download. Sandbox em
                # diretorios da propria sessao pra evitar
                # path traversal.
                file_param = (qs.get("file") or [""])[0]
                if not file_param:
                    self._send(400, "missing 'file' param", "text/plain")
                    return
                target = Path(file_param).resolve()
                # `rdbs` saiu: os uploads agora vao pro cache por conteudo,
                # que e' compartilhado -- deixa-lo entrar no sandbox daria a um
                # visitante o arquivo derivado gerado por outro.
                if not is_within(target, (self.sdir("out"), self.sdir("scd"))):
                    self._send(403, "path outside allowed roots", "text/plain")
                    return
                if not target.is_file():
                    self._send(404, "file not found", "text/plain")
                    return
                data = target.read_bytes()
                ext = target.suffix.lower()
                if ext == ".rdb":
                    ctype = "application/octet-stream"
                elif ext == ".xlsx":
                    ctype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                else:
                    ctype = "application/xml"
                self.send_response(200)
                self.send_header("Content-Type", ctype)
                self.send_header("Content-Length", str(len(data)))
                self.send_header(
                    "Content-Disposition",
                    f'attachment; filename="{target.name}"',
                )
                self.end_headers()
                self.wfile.write(data)
                return
            if path == "/compare":
                relay = (qs.get("relay") or [""])[0]
                ied = (qs.get("ied") or [""])[0]
                gle = (qs.get("gle") or [""])[0]
                with self.session.lock:
                    st = self.sess()
                    rdb = st.rdb
                    scd_path = st.scd_path
                if rdb is None or scd_path is None:
                    self._send(409, "RDB ou SCD nao carregado", "text/plain")
                    return
                entry = rdb_loader.find_gle(rdb, relay, gle)
                if entry is None or not entry.fs_path.is_file():
                    self._send(404, f"GLE {gle!r} nao encontrado para o rele {relay!r}",
                               "text/plain")
                    return
                try:
                    html = _render_compare_page(
                        rdb_relay=relay, ied_name=ied, gle_name=gle,
                        gle_path=entry.fs_path, scd_path=scd_path,
                    )
                except Exception as e:
                    logger.exception("falha renderizando comparacao: %s", e)
                    self._send(500, f"falha: {e}", "text/plain")
                    return
                self._send(200, html, "text/html; charset=utf-8")
                return

            self._send(404, "not found", "text/plain")

        def do_POST(self):
            path = urlparse(self.path).path
            try:
                length = int(self.headers.get("Content-Length", "0"))
            except ValueError:
                length = 0

            if path == "/apply":
                job = self.job()
                job.stage("Aplicando alteracoes", 10)
                body = self.rfile.read(length) if length > 0 else b"{}"
                try:
                    payload = json.loads(body or b"{}")
                    direction = str(payload.get("direction", "")).strip()
                    relay = str(payload.get("relay", "")).strip()
                    ied = str(payload.get("ied", "")).strip()
                    gle = str(payload.get("gle", "")).strip()
                except (json.JSONDecodeError, TypeError):
                    self._send_json(400, {"error": "bad request"})
                    return
                if direction not in ("scd-to-gle", "gle-to-scd"):
                    self._send_json(400, {"error": "direction invalida"})
                    return
                with self.session.lock:
                    st = self.sess()
                    rdb = st.rdb
                    scd_path = st.scd_path
                    # O nome de EXIBICAO, e nao o do arquivo: o acervo guarda
                    # o SCD como `<sha12>.scd`, entao derivar a saida do
                    # caminho batizava o arquivo do usuario de
                    # "72586aeda11e_comments_updated.scd".
                    scd_label = Path(st.scd_name or (scd_path.name if scd_path
                                                     else "arquivo.scd"))
                if rdb is None or scd_path is None:
                    self._send_json(409, {"error": "RDB ou SCD nao carregado"})
                    return
                entry = rdb_loader.find_gle(rdb, relay, gle)
                if entry is None or not entry.fs_path.is_file():
                    self._send_json(404, {"error": f"GLE nao encontrado: {relay}/{gle}"})
                    return
                try:
                    if direction == "scd-to-gle":
                        # O RDB de origem mora no cache por conteudo, que e'
                        # compartilhado; a saida derivada e' desta sessao (e o
                        # /download so serve os diretorios dela).
                        out_path = self.sdir("out") / _with_suffix_before_ext(
                            Path(rdb.display_name), "_comments_updated",
                        ).name
                        result = update_rdb_with_scd_descs(
                            rdb_path=rdb.rdb_path,
                            extract_dir=rdb.extract_dir,
                            relay_name=relay,
                            gle_name=gle,
                            gle_fs_path=entry.fs_path,
                            scd_path=scd_path,
                            ied_name=ied,
                            output_path=out_path,
                            job=self.job(),
                        )
                    else:  # gle-to-scd
                        # O SCD de origem agora tambem pode vir da biblioteca
                        # do projeto (compartilhada com outras ferramentas na
                        # mesma sessao); a saida derivada e' desta ferramenta
                        # (e o /download so serve os diretorios dela).
                        out_path = self.sdir("out") / _with_suffix_before_ext(
                            scd_label, "_comments_updated",
                        ).name
                        result = update_scd_with_gle_comments(
                            scd_path=scd_path,
                            ied_name=ied,
                            gle_path=entry.fs_path,
                            output_path=out_path,
                        )
                except Exception as e:
                    logger.exception("falha aplicando %s: %s", direction, e)
                    self._send_json(500, {"error": str(e)})
                    return
                if result.get("ok"):
                    out_abs = Path(result["output_path"]).resolve()
                    # RDB ou SCD corrigido: e' entrada das outras ferramentas,
                    # entao entra no acervo do projeto em vez de morrer no
                    # link de download desta aba.
                    result["project_file"] = self.publish_output(
                        out_abs, "VB Updater", job=self.job(), logger=logger)
                    result["download_url"] = self.mount_prefix + "/download?file=" + quote(str(out_abs), safe="")
                    result["output_name"] = out_abs.name
                    logger.info("[vb-updater] apply %s -> %s", direction, out_abs.name)
                    self._send_json(200, result)
                else:
                    self._send_json(422, result)
                return

            if path == "/apply-batch":
                body = self.rfile.read(length) if length > 0 else b"{}"
                try:
                    payload = json.loads(body or b"{}")
                    selections = payload.get("selections", [])
                except (json.JSONDecodeError, TypeError):
                    self._send_json(400, {"error": "bad request"})
                    return
                if not isinstance(selections, list) or not selections:
                    self._send_json(400, {"error": "selections vazia"})
                    return
                with self.session.lock:
                    st = self.sess()
                    rdb = st.rdb
                    scd_path = st.scd_path
                if rdb is None or scd_path is None:
                    self._send_json(409, {"error": "RDB ou SCD nao carregado"})
                    return

                resolved: list[dict] = []
                invalid: list[dict] = []
                for sel in selections:
                    if not isinstance(sel, dict):
                        invalid.append({"error": "selecao invalida", "raw": str(sel)})
                        continue
                    relay = str(sel.get("relay", "")).strip()
                    ied = str(sel.get("ied", "")).strip()
                    gle = str(sel.get("gle", "")).strip()
                    if not (relay and ied and gle):
                        invalid.append({
                            "relay": relay, "ied": ied, "gle": gle,
                            "error": "campo obrigatorio ausente",
                        })
                        continue
                    entry = rdb_loader.find_gle(rdb, relay, gle)
                    if entry is None or not entry.fs_path.is_file():
                        invalid.append({
                            "relay": relay, "ied": ied, "gle": gle,
                            "error": "GLE nao encontrado",
                        })
                        continue
                    resolved.append({
                        "relay": relay, "ied": ied,
                        "gle_name": gle, "gle_fs_path": entry.fs_path,
                    })

                if not resolved:
                    self._send_json(422, {
                        "error": "nenhuma selecao valida",
                        "invalid": invalid,
                    })
                    return

                out_path = self.sdir("out") / _with_suffix_before_ext(
                    Path(rdb.display_name), "_batch_comments_updated",
                ).name
                try:
                    result = update_rdb_with_scd_descs_batch(
                        rdb_path=rdb.rdb_path,
                        extract_dir=rdb.extract_dir,
                        scd_path=scd_path,
                        selections=resolved,
                        output_path=out_path,
                        job=self.job(),
                    )
                except Exception as e:
                    logger.exception("falha aplicando batch: %s", e)
                    self._send_json(500, {"error": str(e)})
                    return

                if not result.get("ok"):
                    # Tudo-ou-nada: nenhuma selecao foi gravada, entao nao ha
                    # arquivo pra baixar nem pra publicar no acervo. As
                    # `results` por selecao vao junto pra tela dizer QUAL
                    # falhou.
                    if invalid:
                        result["invalid"] = invalid
                    logger.warning("[vb-updater] apply-batch abortado: %s",
                                   result.get("error"))
                    self._send_json(422, result)
                    return

                out_abs = Path(result["output_path"]).resolve()
                result["project_file"] = self.publish_output(
                    out_abs, "VB Updater", job=self.job(), logger=logger)
                result["download_url"] = self.mount_prefix + "/download?file=" + quote(str(out_abs), safe="")
                result["output_name"] = out_abs.name
                if invalid:
                    result["invalid"] = invalid
                logger.info(
                    "[vb-updater] apply-batch (%d ok, %d fail) -> %s",
                    result.get("succeeded", 0), result.get("failed", 0),
                    out_abs.name,
                )
                self._send_json(200, result)
                return

            if path == "/export-descriptions":
                body = self.rfile.read(length) if length > 0 else b"{}"
                try:
                    payload = json.loads(body or b"{}")
                    selections = payload.get("selections", [])
                except (json.JSONDecodeError, TypeError):
                    self._send_json(400, {"error": "bad request"})
                    return
                if not isinstance(selections, list) or not selections:
                    self._send_json(400, {"error": "selections vazia"})
                    return
                with self.session.lock:
                    st = self.sess()
                    scd_path = st.scd_path
                    scd_label = Path(st.scd_name or (scd_path.name if scd_path
                                                     else "arquivo.scd"))
                if scd_path is None:
                    self._send_json(409, {"error": "SCD nao carregado"})
                    return
                # Coleta IEDs preservando ordem e deduplicando. Tambem mapeia
                # IED -> relay name (RDB) pra mostrar na aba.
                ied_names: list[str] = []
                seen: set[str] = set()
                rdb_by_ied: dict[str, str] = {}
                for sel in selections:
                    if not isinstance(sel, dict):
                        continue
                    ied = str(sel.get("ied", "")).strip()
                    relay = str(sel.get("relay", "")).strip()
                    if not ied or ied in seen:
                        continue
                    seen.add(ied)
                    ied_names.append(ied)
                    if relay:
                        rdb_by_ied[ied] = relay
                if not ied_names:
                    self._send_json(400, {"error": "nenhum IED valido na selecao"})
                    return
                try:
                    xlsx_bytes = build_vb_descriptions_xlsx(
                        scd_path=scd_path, ied_names=ied_names,
                        rdb_by_ied=rdb_by_ied,
                    )
                except Exception as e:
                    logger.exception("falha gerando xlsx: %s", e)
                    self._send_json(500, {"error": str(e)})
                    return
                # Salva no diretorio de SCD da sessao (mesmo sandbox de /download).
                out_name = _with_suffix_before_ext(
                    scd_label, "_descriptions",
                ).with_suffix(".xlsx").name
                out_path = (self.sdir("scd") / out_name).resolve()
                try:
                    out_path.write_bytes(xlsx_bytes)
                except OSError as e:
                    self._send_json(500, {"error": f"falha ao salvar xlsx: {e}"})
                    return
                logger.info(
                    "[vb-updater] export-descriptions: %d IEDs -> %s",
                    len(ied_names), out_name,
                )
                self._send_json(200, {
                    "ok": True,
                    "output_name": out_name,
                    "project_file": self.publish_output(
                        out_path, "VB Updater", job=self.job(), logger=logger),
                    "download_url": self.mount_prefix + "/download?file=" + quote(str(out_path), safe=""),
                    "ied_count": len(ied_names),
                })
                return

            if path == "/import-descriptions":
                if length <= 0:
                    self._send_json(400, {"error": "empty upload"})
                    return
                if length > _SCD_MAX_BYTES:
                    self._send_json(413, {"error": "arquivo grande demais"})
                    return
                with self.session.lock:
                    st = self.sess()
                    scd_path = st.scd_path
                    scd_label = Path(st.scd_name or (scd_path.name if scd_path
                                                     else "arquivo.scd"))
                if scd_path is None:
                    self._send_json(409, {"error": "SCD nao carregado"})
                    return
                # IEDs selecionados (filtro opcional): aceita header
                # X-Selected-Ieds com uma lista JSON (URL-encoded). E' JSON e
                # nao "a,b" porque o encodeURIComponent do cliente escapa a
                # virgula literal de um nome exatamente como o separador: depois
                # do unquote as duas ficam indistinguiveis e um IED chamado
                # "A,B" viraria dois.
                raw_sel = self.headers.get("X-Selected-Ieds", "")
                selected_ieds: set[str] | None = None
                if raw_sel:
                    try:
                        parsed_sel = json.loads(unquote(raw_sel))
                        if isinstance(parsed_sel, list):
                            selected_ieds = {
                                s.strip() for s in parsed_sel
                                if isinstance(s, str) and s.strip()
                            }
                    except Exception:
                        selected_ieds = None
                xlsx_bytes = self.rfile.read(length)
                try:
                    parsed = parse_vb_descriptions_xlsx(xlsx_bytes)
                except Exception as e:
                    logger.exception("falha parseando xlsx: %s", e)
                    self._send_json(422, {"error": f"xlsx invalido: {e}"})
                    return
                if not parsed:
                    self._send_json(422, {
                        "error": "nenhuma descricao valida encontrada no xlsx "
                                 "(verifique que A1='IED:' e B1=<nome do IED>)",
                    })
                    return
                # Filtra pelos selecionados, se houver.
                ignored: list[str] = []
                if selected_ieds is not None:
                    filtered = {}
                    for ied, vb_map in parsed.items():
                        if ied in selected_ieds:
                            filtered[ied] = vb_map
                        else:
                            ignored.append(ied)
                    parsed = filtered
                if not parsed:
                    self._send_json(422, {
                        "error": "nenhum IED do xlsx esta na selecao atual",
                        "ignored_ieds": ignored,
                    })
                    return
                out_path = self.sdir("out") / _with_suffix_before_ext(
                    scd_label, "_descriptions_imported").name
                try:
                    result = update_scd_with_descriptions_multi(
                        scd_path=scd_path,
                        descriptions_by_ied=parsed,
                        output_path=out_path,
                    )
                except Exception as e:
                    logger.exception("falha aplicando descricoes: %s", e)
                    self._send_json(500, {"error": str(e)})
                    return
                out_abs = Path(result["output_path"]).resolve()
                result["project_file"] = self.publish_output(
                    out_abs, "VB Updater", job=self.job(), logger=logger)
                result["download_url"] = self.mount_prefix + "/download?file=" + quote(str(out_abs), safe="")
                result["output_name"] = out_abs.name
                if ignored:
                    result["ignored_ieds"] = ignored
                logger.info(
                    "[vb-updater] import-descriptions: %d IEDs aplicados -> %s",
                    len(parsed), out_abs.name,
                )
                self._send_json(200, result)
                return

            if path in ("/select-rdb", "/select-scd"):
                # O corpo do antigo /rdb-upload ou /scd-upload, do
                # `process_upload`/`load_scd` pra frente: o arquivo ja foi
                # recebido e extraido/validado em /files/.
                want = (filelib.KIND_RDB if path == "/select-rdb"
                        else filelib.KIND_SCD)
                body = self._read_json_body()
                sha = (body.get("sha256") or "").strip()
                lib = filelib.library_for(sessions, self.session)
                with self.session.lock:
                    entry = lib.get(sha)
                if entry is None or entry.kind != want:
                    self._send_json(404, {
                        "error": "Arquivo não está mais no projeto."})
                    return
                job = self.job()
                with self.session.lock:
                    st = self.sess()
                    if want == filelib.KIND_RDB:
                        st.rdb = entry.rdb
                        logger.info("[vb-updater] RDB '%s' (%s) escolhido; "
                                    "%d relé(s) com GLE",
                                    entry.display_name, entry.short_sha,
                                    len(entry.rdb.relays))
                    else:
                        st.scd_path = entry.scd_path
                        st.scd_name = entry.display_name
                        logger.info("[vb-updater] SCD '%s' (%s) escolhido",
                                    entry.display_name, entry.short_sha)
                    # O cruzamento RDB x SCD so acontece quando os dois
                    # existem; `_maybe_match` ja sabe disso.
                    job.stage("Cruzando RDB com SCD", 60)
                    _maybe_match(st)
                job.finish("Arquivo carregado")
                self._send_json(200, _state_payload(self.sess()))
                return

            self._send(404, "not found", "text/plain")

    return Handler
