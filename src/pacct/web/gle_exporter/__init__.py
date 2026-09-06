"""
GLE Variable Comment Exporter: extracts the list of SYMBOL instances of each
GLE in an RDB and exports it as Excel for bulk editing. The user edits the
port (input/output) comments and reimports the Excel to produce an updated
RDB.

Flow:
  1. The user uploads an RDB.
  2. The app lists the RDB's relays with a per-relay GLE selection.
  3. The "Exportar Excel" button builds an xlsx with one sheet per selected
     (relay, GLE), holding every SYMBOL instance of that GLE.
  4. The user edits the "Input Comment"/"Output Comment" cells.
  5. The "Importar Excel" button applies the changes to the RDB and
     downloads a new RDB.

Each SYMBOL in the GLE is an "instance" identified by (page, element_id). The
same name (e.g. TMB1A) can appear several times on different pages; each
occurrence is one row in the Excel.

    templates/  landing.html
"""

from __future__ import annotations

import json
import logging
import re
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any
from urllib.parse import quote, urlparse

from sellib import rdb as rdb_loader
from sellib.gle import parse_gle
from sellib.rdb import RdbInfo

from pacct.paths import GLE_EXPORTER_TEMPLATES_DIR, is_within
from pacct.web import rdb_write
from pacct.web.project_files import library as filelib
from pacct.web.rdb_write import (
    resolve_gle_stream_path as _resolve_gle_stream_path,
)
from pacct.web.rdb_write import (
    with_suffix_before_ext as _with_suffix_before_ext,
)
from pacct.web.rdb_write import (
    xml_text_escape,
)
from pacct.web.session import SessionHandler
from pacct.web.xlsx_names import sanitize_sheet_name as _sanitize_sheet_name

_logger = logging.getLogger(__name__)


def load_template(name: str) -> str:
    """Read one template. Read at import time, like the GLV and the DNP map."""
    return (GLE_EXPORTER_TEMPLATES_DIR / name).read_text(encoding="utf-8")


_XLSX_MAX_BYTES = 50 * 1024 * 1024


# Markers: they identify an xlsx as coming from this tool's export.
_XLSX_RELAY_MARKER = "Relay:"
_XLSX_GLE_MARKER = "GLE:"
# Row 4 header (each column = one cell in this header).
_XLSX_HEADERS = (
    "Page", "Element ID", "Type", "Variable", "Side", "Port", "Label", "Comment",
)
# 1-based column of each field (stable for the parser and the tests).
_COL_PAGE, _COL_EID, _COL_TYPE, _COL_VAR, _COL_SIDE, _COL_PORT, _COL_LABEL, _COL_CMT = range(1, 9)

# Output directory for the xlsx (reuses the /download sandbox).
# Uploads and outputs live in cache/sessions/<sid>/ (see pacct.web.session);
# there is no directory shared between users any more.


# -----------------------------------------------------------------------------
# Extractor: GLE.xml -> [PortInstance]
# -----------------------------------------------------------------------------

# GLE XML types with a `physical_instance_name` that we want to export:
#   - SYMBOL (IO variables, Relay Word bits): 1 in + 1 out, no fixed labels.
#   - 4xx/7xx stateful blocks: PLT, ALT, PCNDTIMER, PCN, AST, PSV, LATCH,
#     TIMER, COUNTER. Port labels come from the relay_model JSON.
# Pure gates (AND/OR/NOT/EQ/...) are left out -- they have no editable
# instance name and their port comments carry no commissioning meaning.
_EXPORTABLE_XML_TYPES = frozenset({
    "SYMBOL",
    "PLT", "ALT", "PCNDTIMER", "PCN", "AST", "PSV",
    "LATCH", "TIMER", "COUNTER",
})


@dataclass(frozen=True)
class PortInstance:
    """One concrete port (in a GLE element) with its fixed label and the
    user's free-form comment."""
    page: str
    element_id: str
    xml_type: str        # type of the <element>: SYMBOL, PLT, LATCH, ...
    name: str            # physical_instance_name (ex.: VB203, _PLT06, TMB1A)
    side: str            # "input" or "output"
    port_index: int      # the <port>'s `index` attribute (0, 1, 2, ...)
    label: str           # fixed pin label (S/R/Q/in/PU/...) or "" if absent
    comment: str         # free text from the user; "" if <comment/> is empty


def extract_port_instances_from_gle(
    gle_path: Path, relay_model=None,
) -> list[PortInstance]:
    """Read a GLE.xml and return a list of PortInstance, one per port of
    each exportable element.

    GLE convention: inside each `<logic_element>` there are (in order) two
    `<ports>` blocks -- the 1st is the input side (left), the 2nd is the
    output side (right). Each `<port>` has a numeric `index` (0, 1, 2).

    If `relay_model` is given, the pin labels (S/R/Q/...) are resolved via
    `relay_model.port_label`. Without the model, label stays "".
    """
    out: list[PortInstance] = []
    try:
        root = parse_gle(gle_path)
    except (OSError, ET.ParseError, UnicodeDecodeError) as e:
        _logger.warning("erro lendo GLE %s: %s", gle_path, e)
        return out

    for page in root.iter("page"):
        page_name = (page.attrib.get("name") or "").strip()
        for el in page.iter("element"):
            xml_type = (el.attrib.get("type") or "").strip()
            if xml_type not in _EXPORTABLE_XML_TYPES:
                continue
            element_id = (el.attrib.get("id") or "").strip()
            le = el.find("logic_element")
            if le is None:
                continue
            name = (le.attrib.get("physical_instance_name") or "").strip()
            if not name:
                continue
            port_groups = le.findall("ports")
            for grp_idx, pg in enumerate(port_groups[:2]):
                side = "input" if grp_idx == 0 else "output"
                for port_el in pg.findall("port"):
                    try:
                        idx = int(port_el.attrib.get("index", "0"))
                    except ValueError:
                        idx = 0
                    cmt_el = port_el.find("comment")
                    cmt = (cmt_el.text or "").strip() if cmt_el is not None else ""
                    label = ""
                    if relay_model is not None:
                        lbl = relay_model.port_label(xml_type, side, idx)
                        if lbl is not None:
                            label = lbl
                    out.append(PortInstance(
                        page=page_name, element_id=element_id,
                        xml_type=xml_type, name=name, side=side,
                        port_index=idx, label=label, comment=cmt,
                    ))
    return out


# -----------------------------------------------------------------------------
# Writer: edits the GLE comments in bytes (keeps the OLE stream size)
# -----------------------------------------------------------------------------

# Matches an <element id="N" type="TIPO" ...> ... </element>. We use named
# groups so as not to depend on indices (adding `type` shifts the numeric
# ones, which has already been a source of bugs).
# Non-greedy on the body: `<element>` does not nest, and `</logic_element>`
# (which is INSIDE) does not match a literal `</element>` -- so the first
# `</element>` found is always the outer block's.
_GLE_ELEMENT_RE = re.compile(
    rb'(?P<open><element\s+id="(?P<id>\d+)"\s+type="(?P<type>[^"]+)"[^>]*>)'
    rb'(?P<body>.*?)'
    rb'(?P<close></element>)',
    re.DOTALL,
)

# Matches a <ports>...</ports> block OR <ports/> (self-closing) inside the
# logic_element. Also non-greedy: `<ports>` does not nest.
_PORTS_BLOCK_RE = re.compile(
    rb'<ports\s*/>|<ports\b[^>]*>.*?</ports>',
    re.DOTALL,
)

# Inside a <ports>...</ports>, matches the comment of <port index="N"> (N
# arrives by substituting {idx}). Two comment shapes in the GLE:
#   <comment />                  (self-closing, empty)
#   <comment>TEXT</comment>      (with or without text)
# group(1) = the port opening + the opening of <comment;
# group(2) = the comment body;
# group(3) = the port closing.
def _port_by_index_re(idx: int) -> re.Pattern[bytes]:
    pat = (
        rb'(<port\s+index="' + str(int(idx)).encode("ascii") + rb'"[^>]*>\s*<comment)'
        rb'(\s*/>|\s*>[^<]*</comment>)'
        rb'(\s*</port>)'
    )
    return re.compile(pat, re.DOTALL)


def _build_comment_node(new_comment: str) -> bytes:
    """Build the new <comment...> node as latin-1 bytes.
    Empty -> self-closing ` />`; non-empty -> `>TEXT</comment>`.
    """
    if not new_comment:
        return b" />"
    text_bytes = xml_text_escape(new_comment).encode("latin-1", errors="replace")
    return b">" + text_bytes + b"</comment>"


def _set_port_comment(
    ports_block: bytes, port_index: int, new_comment: str,
) -> tuple[bytes, bool]:
    """In a <ports>...</ports>, replace the <port index=port_index> comment.
    Returns (new_bytes, was_updated). False when the block is self-closing
    (no port) or has no port with that index."""
    replacement = _build_comment_node(new_comment)
    pat = _port_by_index_re(port_index)

    def _sub(m: re.Match) -> bytes:
        return m.group(1) + replacement + m.group(3)

    new_block, n = pat.subn(_sub, ports_block, count=1)
    return new_block, n > 0


# updates: { element_id -> { (side, port_index) -> new_comment } }
PortUpdates = dict[str, dict[tuple[str, int], str]]


def update_port_comments_in_gle_bytes(
    raw: bytes, updates: PortUpdates,
) -> tuple[bytes, dict]:
    """Update the comments of specific ports by (element_id, side,
    port_index) in a GLE.

    `updates`: { element_id -> { (side, port_index) -> new_comment } }.
    `side` ∈ {"input","output"}. Empty string = force <comment /> self-closing.

    Returns (new_bytes, stats). stats has the keys:
      - elements_touched: # of elements with at least one port touched
      - ports_updated:    # total of port-comments replaced
      - ports_skipped:    update requests that did not find the expected port
      - elements_missing: ids in `updates` that were not found in the GLE
    """
    stats = {"elements_touched": 0, "ports_updated": 0,
             "ports_skipped": 0, "elements_missing": 0}
    seen_ids: set[str] = set()

    def replace_element(m: re.Match) -> bytes:
        eid = m.group("id").decode("ascii")
        if eid not in updates:
            return m.group(0)
        seen_ids.add(eid)
        upd = updates[eid]
        head = m.group("open")
        body = m.group("body")
        tail = m.group("close")

        # Locate the <ports>...</ports> blocks (at most 2: input and output).
        port_matches = list(_PORTS_BLOCK_RE.finditer(body))
        # Group updates by side to process them block-by-block.
        by_side: dict[str, dict[int, str]] = {"input": {}, "output": {}}
        for (side, idx), new_cmt in upd.items():
            if side in by_side:
                by_side[side][idx] = new_cmt

        out_parts: list[bytes] = []
        last_end = 0
        touched_here = 0
        for i, pm in enumerate(port_matches[:2]):
            out_parts.append(body[last_end:pm.start()])
            side = "input" if i == 0 else "output"
            side_updates = by_side.get(side, {})
            if not side_updates:
                out_parts.append(pm.group(0))
            else:
                new_block = pm.group(0)
                for idx, new_cmt in side_updates.items():
                    new_block, ok = _set_port_comment(new_block, idx, new_cmt)
                    if ok:
                        stats["ports_updated"] += 1
                        touched_here += 1
                    else:
                        stats["ports_skipped"] += 1
                out_parts.append(new_block)
            last_end = pm.end()
        out_parts.append(body[last_end:])

        # Updates aimed at a side whose <ports> does not exist -> skipped.
        for i in range(len(port_matches), 2):
            side = "input" if i == 0 else "output"
            stats["ports_skipped"] += len(by_side.get(side, {}))

        if touched_here:
            stats["elements_touched"] += 1
        return head + b"".join(out_parts) + tail

    new_raw = _GLE_ELEMENT_RE.sub(replace_element, raw)
    stats["elements_missing"] = len(set(updates.keys()) - seen_ids)
    return new_raw, stats


# -----------------------------------------------------------------------------
# Fit XML to a target byte size (same pattern as the vb_updater)
# -----------------------------------------------------------------------------
# Excel I/O
# -----------------------------------------------------------------------------

def build_xlsx_for_selections(
    selections: list[dict],
) -> bytes:
    """Build an .xlsx (bytes) with one sheet per selection.

    Each `selection` is a dict {relay, gle, gle_path, relay_model}. The sheet
    holds:
      A1: "Relay:"      B1: <relay_name>
      A2: "GLE:"        B2: <gle_name>
      Row 3: blank
      Row 4: headers [Page, Element ID, Type, Variable, Side, Port, Label, Comment]
      Row 5+: one row per port of each exportable element.

    `relay_model` (optional) enables the fixed port labels (S/R/Q/...);
    without it the Label column stays empty but the table is still valid.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    # A fresh `Workbook()` always has one sheet, but `active` is
    # Optional in the stubs -- every sheet here is added by name.
    default = wb.active
    if default is not None:
        wb.remove(default)

    used: set[str] = set()
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="FF1F6FEB")
    meta_font = Font(bold=True)

    for sel in selections:
        relay = sel["relay"]
        gle = sel["gle"]
        gle_path = sel["gle_path"]
        relay_model = sel.get("relay_model")
        ports = extract_port_instances_from_gle(gle_path, relay_model=relay_model)

        sheet_base = f"{relay} {gle}".strip()
        ws = wb.create_sheet(title=_sanitize_sheet_name(sheet_base, used))
        ws["A1"] = _XLSX_RELAY_MARKER
        ws["B1"] = relay
        ws["A1"].font = meta_font
        ws["A2"] = _XLSX_GLE_MARKER
        ws["B2"] = gle
        ws["A2"].font = meta_font

        for col, val in enumerate(_XLSX_HEADERS, start=1):
            c = ws.cell(row=4, column=col, value=val)
            c.font = header_font
            c.fill = header_fill
            c.alignment = Alignment(horizontal="left")

        for i, p in enumerate(ports, start=5):
            ws.cell(row=i, column=_COL_PAGE,  value=p.page)
            ws.cell(row=i, column=_COL_EID,   value=p.element_id)
            ws.cell(row=i, column=_COL_TYPE,  value=p.xml_type)
            ws.cell(row=i, column=_COL_VAR,   value=p.name)
            ws.cell(row=i, column=_COL_SIDE,  value=p.side)
            ws.cell(row=i, column=_COL_PORT,  value=p.port_index)
            ws.cell(row=i, column=_COL_LABEL, value=p.label)
            ws.cell(row=i, column=_COL_CMT,   value=p.comment)

        # Widths: sized roughly so nothing shows up cut off.
        widths = {
            _COL_PAGE: 28, _COL_EID: 10, _COL_TYPE: 12, _COL_VAR: 18,
            _COL_SIDE: 8,  _COL_PORT: 6, _COL_LABEL: 10, _COL_CMT: 60,
        }
        for col_idx, w in widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = w
        ws.freeze_panes = "A5"

    if not wb.sheetnames:
        wb.create_sheet(title="empty")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Type returned by the xlsx parser:
#   {(relay, gle) -> {element_id -> {(side, port_index) -> new_comment}}}
XlsxUpdates = dict[tuple[str, str], PortUpdates]


def parse_xlsx_to_updates(xlsx_bytes: bytes) -> XlsxUpdates:
    """Read an xlsx built by `build_xlsx_for_selections` (possibly edited)
    and return the port-by-port updates structure.

    Rules:
      - A sheet must have A1='Relay:' and A2='GLE:' (the markers). Other
        sheets are silently ignored.
      - Rows with an empty Element ID or an unknown Side are skipped.
      - If the Comment cell is None (cell erased/never filled in the
        original export), it does NOT enter updates -> the port is untouched.
      - An explicit "" Comment = a request to clear it (becomes <comment />).
      - Side is normalised to "input"/"output" (case insensitive).
    """
    from openpyxl import load_workbook

    wb = load_workbook(filename=BytesIO(xlsx_bytes), data_only=True, read_only=True)
    out: XlsxUpdates = {}
    try:
        for ws in wb.worksheets:
            a1 = ws.cell(row=1, column=1).value
            a2 = ws.cell(row=2, column=1).value
            if not (isinstance(a1, str) and a1.strip() == _XLSX_RELAY_MARKER):
                continue
            if not (isinstance(a2, str) and a2.strip() == _XLSX_GLE_MARKER):
                continue
            relay = ws.cell(row=1, column=2).value
            gle = ws.cell(row=2, column=2).value
            if not relay or not gle:
                continue
            relay = str(relay).strip()
            gle = str(gle).strip()
            per_gle: PortUpdates = out.setdefault((relay, gle), {})
            for row in ws.iter_rows(min_row=5, max_col=len(_XLSX_HEADERS),
                                     values_only=True):
                if row is None:
                    continue
                # Cells out of a spreadsheet the user uploaded: any of
                # openpyxl's value types, or None for a short row.
                # Every read below either `str()`s it or converts
                # inside a try, which is what makes that safe.
                cells: list[Any] = list(row) + [None] * len(_XLSX_HEADERS)
                _page = cells[_COL_PAGE - 1]
                eid = cells[_COL_EID - 1]
                _xml_type = cells[_COL_TYPE - 1]
                _var = cells[_COL_VAR - 1]
                side_raw = cells[_COL_SIDE - 1]
                port_raw = cells[_COL_PORT - 1]
                _label = cells[_COL_LABEL - 1]
                cmt = cells[_COL_CMT - 1]
                if eid is None:
                    continue
                eid_str = str(eid).strip()
                if not eid_str.isdigit():
                    continue
                side = str(side_raw or "").strip().lower()
                if side not in ("input", "output"):
                    continue
                try:
                    port_idx = int(port_raw) if port_raw is not None else 0
                except (TypeError, ValueError):
                    continue
                if cmt is None:
                    continue  # empty cell (not edited) -> do not touch
                entry = per_gle.setdefault(eid_str, {})
                entry[(side, port_idx)] = str(cmt).strip()
    finally:
        wb.close()
    return {k: v for k, v in out.items() if v}


def diff_updates_against_gle(
    gle_path: Path, updates: PortUpdates, relay_model=None,
) -> PortUpdates:
    """Filter updates, dropping (side, port_index) already equal in the GLE."""
    if not updates:
        return {}
    current: dict[str, dict[tuple[str, int], str]] = {}
    for p in extract_port_instances_from_gle(gle_path, relay_model=relay_model):
        current.setdefault(p.element_id, {})[(p.side, p.port_index)] = p.comment

    out: PortUpdates = {}
    for eid, ports in updates.items():
        if eid not in current:
            # element_id does not exist in the current GLE -- kept so the
            # writer can count it as missing.
            out[eid] = dict(ports)
            continue
        cur = current[eid]
        delta: dict[tuple[str, int], str] = {}
        for key, new_cmt in ports.items():
            if cur.get(key, "") != new_cmt:
                delta[key] = new_cmt
        if delta:
            out[eid] = delta
    return out


# -----------------------------------------------------------------------------
# Orchestrator: applies the xlsx edits to the RDB and produces a new RDB
# -----------------------------------------------------------------------------

def apply_xlsx_updates_to_rdb(
    *,
    rdb_info: RdbInfo,
    xlsx_updates: XlsxUpdates,
    output_path: Path,
    job=None,
) -> dict:
    """Apply the xlsx changes to the RDB and produce `output_path`.

    `xlsx_updates`: (relay, gle) -> {element_id: {(side, port_idx): comment}}.

    Two passes, and the order matters. The FIRST one only reads: it resolves
    each (relay, GLE) to its stream in the OLE and computes the new bytes,
    without touching disk. Only if ALL of them succeed does the second pass
    write, in one go, via `rdb_write.write_streams` -- which uses
    `write_stream` when every stream keeps its size and rebuilds the
    container when one does not, always atomically onto the destination.

    It is all-or-nothing on purpose. Before, each selection was written into
    the output RDB as soon as it was ready: if the third failed, the first
    two were already in the file, the response still said `ok: True`, and
    that half-applied RDB went into the project library just like a complete
    one. A half-written settings file is indistinguishable from a whole one
    once it leaves here.

    Returns {ok, output_path, results, succeeded, failed, totals, method}.
    With `ok: False`, NO file was written and `output_path` does not exist.
    """
    from sellib.models import relay_models as _rm

    # RelayModel cache by name (resolved only once per relay).
    model_cache: dict[str, object] = {}
    def _model_for(relay_name: str):
        if relay_name in model_cache:
            return model_cache[relay_name]
        entry_relay = next(
            (r for r in rdb_info.relays if r.name == relay_name), None,
        )
        m = _rm.lookup(entry_relay.model) if (entry_relay and entry_relay.model) else None
        model_cache[relay_name] = m
        return m

    results: list[dict] = []
    totals = {"elements_touched": 0, "ports_updated": 0,
              "ports_skipped": 0, "elements_missing": 0}
    streams: dict[tuple[str, ...], bytes] = {}

    # -- pass 1: read only ----------------------------------------------------
    if job:
        job.stage("Conferindo os GLEs alterados", 10)
    for (relay, gle), elem_updates in xlsx_updates.items():
        # One row of `results`: heterogeneous by design -- the flag, the
        # stats dict, the byte count and the reason all live in it.
        entry: dict[str, Any] = {"relay": relay, "gle": gle}
        gle_entry = rdb_loader.find_gle(rdb_info, relay, gle)
        if gle_entry is None or not gle_entry.fs_path.is_file():
            entry["ok"] = False
            entry["error"] = "GLE nao encontrado no RDB"
            results.append(entry)
            continue

        # Filter out updates that already match the current content.
        delta = diff_updates_against_gle(
            gle_entry.fs_path, elem_updates,
            relay_model=_model_for(relay),
        )
        if not delta:
            entry["ok"] = True
            entry["stats"] = {"elements_touched": 0, "ports_updated": 0,
                              "ports_skipped": 0, "elements_missing": 0}
            entry["note"] = "nada a aplicar (valores ja batem)"
            results.append(entry)
            continue

        stream_parts = _resolve_gle_stream_path(rdb_info.extract_dir,
                                                gle_entry.fs_path)
        if not stream_parts:
            entry["ok"] = False
            entry["error"] = f"stream nao localizado no OLE: {gle_entry.rel_path}"
            results.append(entry)
            continue

        # The read comes from the EXTRACTED file, not from the OLE: it is
        # the same content, and this way pass 1 never opens the RDB.
        original = gle_entry.fs_path.read_bytes()
        updated, stats = update_port_comments_in_gle_bytes(original, delta)
        streams[tuple(stream_parts)] = updated
        entry["ok"] = True
        entry["stats"] = stats
        entry["original_stream_bytes"] = len(original)
        for k in totals:
            totals[k] += stats.get(k, 0)
        results.append(entry)

    succeeded = sum(1 for r in results if r.get("ok"))
    failed = len(results) - succeeded
    if failed:
        return {
            "ok": False,
            "error": (f"{failed} de {len(results)} seleção(ões) falharam; "
                      "nenhum RDB foi gravado."),
            "results": results,
            "succeeded": succeeded,
            "failed": failed,
            "totals": totals,
        }

    # -- pass 2: one write ---------------------------------------------------
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        if streams:
            method = rdb_write.write_streams(rdb_info.rdb_path, output_path,
                                             streams, job=job)
        else:
            # Every selection was a no-op. The user still gets the RDB
            # (which is what happened before, when the copy came first),
            # only with no stream rewritten.
            rdb_write.copy_only(rdb_info.rdb_path, output_path)
            method = "copy"
    except rdb_write.RdbWriteError as e:
        return {
            "ok": False,
            "error": str(e),
            "results": results,
            "succeeded": succeeded,
            "failed": failed,
            "totals": totals,
        }

    return {
        "ok": True,
        "output_path": str(output_path),
        "method": method,
        "results": results,
        "succeeded": succeeded,
        "failed": failed,
        "totals": totals,
    }


# -----------------------------------------------------------------------------
# Session state
# -----------------------------------------------------------------------------

@dataclass
class _SessionState:
    rdb: RdbInfo | None = None


def _state_payload(st: _SessionState) -> dict:
    if st.rdb is None:
        return {
            "has_rdb": False,
            "rdb_name": None,
            "relays": [],
        }
    return {
        "has_rdb": True,
        "rdb_name": st.rdb.display_name,
        "relays": [
            {
                "name": r.name,
                "model": r.model,
                "ip": r.ip,
                "gles": [{"name": g.name, "filename": g.filename}
                         for g in r.gles],
            }
            for r in st.rdb.relays
        ],
    }


# -----------------------------------------------------------------------------
# HTML
# -----------------------------------------------------------------------------

LANDING_HTML = load_template("landing.html")

# The numbered navigation is the same on the nine screens -- lives in theme.py.


# -----------------------------------------------------------------------------
# Server
# -----------------------------------------------------------------------------

def build_gle_exporter_handler(logger: logging.Logger, sessions) -> type:
    """Return the GLE Variable Comment Exporter handler class.

    Opens no server: serving is done by the single dispatcher of
    `pacct.web.mount`, which mounts this handler at `/gle-exporter/`. State
    and files are per session (`self.sess()` / `self.sdir()`), not per
    process.
    """

    class Handler(SessionHandler):
        session_key = "gle-exporter"
        state_factory = _SessionState
        server_sessions = sessions

        def do_GET(self):
            parsed = urlparse(self.path)
            path = parsed.path
            if path in ("/", "/index.html"):
                self._send(200, LANDING_HTML, "text/html; charset=utf-8")
                return
            if path == "/gle-state":
                self._send_json(200, {"ok": True})
                return
            if path == "/state":
                self._send_json(200, _state_payload(self.sess()))
                return
            if path == "/download":
                from urllib.parse import parse_qs as _parse_qs
                qs = _parse_qs(parsed.query)
                file_param = (qs.get("file") or [""])[0]
                if not file_param:
                    self._send(400, "missing 'file' param", "text/plain")
                    return
                target = Path(file_param).resolve()
                # `rdbs` is gone: uploads now go to the content cache,
                # which is shared -- letting it into the sandbox would hand
                # one visitor the derived file generated by another.
                if not is_within(target, (self.sdir("out"), self.sdir("xlsx"))):
                    self._send(403, "path outside allowed roots", "text/plain")
                    return
                if not target.is_file():
                    self._send(404, "file not found", "text/plain")
                    return
                ext = target.suffix.lower()
                if ext == ".xlsx":
                    ctype = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                else:
                    ctype = "application/octet-stream"
                # Streamed: an exported RDB is 40-140 MB, and `read_bytes()`
                # held all of it per concurrent download on a threaded server.
                # `send_file` also writes the RFC 5987 name -- these names come
                # from a DISPLAY name and carry accents, and `send_header`
                # encodes latin-1 strict, so a plain `filename="..."` with an
                # en-dash raises after the status line has gone out.
                self.send_file(target, ctype, download_name=target.name)
                return

            self._send(404, "not found", "text/plain")

        def do_POST(self):
            path = urlparse(self.path).path
            try:
                length = int(self.headers.get("Content-Length", "0"))
            except ValueError:
                length = 0

            if path == "/select-rdb":
                # The body of the old /rdb-upload, from `process_upload`
                # onward: the file was already received and extracted in
                # /files/.
                payload = self._read_json_body()
                sha = (payload.get("sha256") or "").strip()
                lib = filelib.library_for(sessions, self.require_session())
                with self.require_session().lock:
                    chosen = lib.get(sha)
                if chosen is None or chosen.kind != filelib.KIND_RDB:
                    self._send_json(404, {
                        "error": "Arquivo não está mais no projeto."})
                    return
                info = chosen.require_rdb()
                logger.info("[gle-exporter] RDB '%s' (%s) escolhido; "
                            "%d relé(s) com GLE",
                            info.display_name, info.sha256[:16],
                            len(info.relays))
                st = self.sess()
                with self.require_session().lock:
                    st.rdb = info
                self._send_json(200, _state_payload(self.sess()))
                return

            if path == "/export":
                job = self.job()
                job.stage("Lendo selecao", 0)
                body = self.rfile.read(length) if length > 0 else b"{}"
                try:
                    payload = json.loads(body or b"{}")
                    selections = payload.get("selections", [])
                except (json.JSONDecodeError, TypeError):
                    self._send_json(400, {"error": "bad request"})
                    return
                if not isinstance(selections, list) or not selections:
                    self._send_json(400, {"error": "selections vazia"})
                    return
                with self.require_session().lock:
                    rdb = self.sess().rdb
                if rdb is None:
                    self._send_json(409, {"error": "RDB nao carregado"})
                    return
                from sellib.models import relay_models as _rm
                model_cache: dict[str, object] = {}
                def _model_for(relay_name: str):
                    if relay_name in model_cache:
                        return model_cache[relay_name]
                    r_entry = next(
                        (r for r in rdb.relays if r.name == relay_name), None,
                    )
                    m = (_rm.lookup(r_entry.model)
                         if (r_entry and r_entry.model) else None)
                    model_cache[relay_name] = m
                    return m

                resolved: list[dict] = []
                for sel in selections:
                    if not isinstance(sel, dict):
                        continue
                    relay = str(sel.get("relay", "")).strip()
                    gle = str(sel.get("gle", "")).strip()
                    if not (relay and gle):
                        continue
                    gle_entry = rdb_loader.find_gle(rdb, relay, gle)
                    if gle_entry is None or not gle_entry.fs_path.is_file():
                        continue
                    resolved.append({
                        "relay": relay, "gle": gle,
                        "gle_path": gle_entry.fs_path,
                        "relay_model": _model_for(relay),
                    })
                if not resolved:
                    self._send_json(422, {"error": "nenhuma selecao valida"})
                    return
                try:
                    xlsx_bytes = build_xlsx_for_selections(resolved)
                except Exception as e:
                    logger.exception("falha gerando xlsx: %s", e)
                    self._send_json(500, {"error": str(e)})
                    return
                out_name = _with_suffix_before_ext(
                    Path(rdb.display_name), "_gle_comments",
                ).with_suffix(".xlsx").name
                job.stage("Gravando planilha", 90)
                out_path = (self.sdir("xlsx") / out_name).resolve()
                try:
                    out_path.write_bytes(xlsx_bytes)
                except OSError as e:
                    self._send_json(500, {"error": f"falha ao salvar xlsx: {e}"})
                    return
                # Count the exported ports to show in the summary.
                total_ports = 0
                for sel in resolved:
                    total_ports += len(extract_port_instances_from_gle(
                        sel["gle_path"], relay_model=sel.get("relay_model"),
                    ))
                logger.info(
                    "[gle-exporter] export: %d aba(s), %d porta(s) -> %s",
                    len(resolved), total_ports, out_name,
                )
                # The spreadsheet enters the library too: no tool selects
                # it (nobody asks for `?kind=xlsx`), but this way it can be
                # downloaded again later from one single place, instead of
                # only from the tab that happened to generate it.
                project = self.publish_output(out_path, "GLE Exporter",
                                              job=job, logger=logger)
                job.finish("Planilha gerada")
                self._send_json(200, {
                    "ok": True,
                    "output_name": out_name,
                    "project_file": project,
                    "download_url": self.mount_prefix + "/download?file=" + quote(str(out_path), safe=""),
                    "selections_count": len(resolved),
                    "total_ports": total_ports,
                })
                return

            if path == "/import":
                job = self.job()
                job.stage("Recebendo planilha", 0)
                if length <= 0:
                    self._send_json(400, {"error": "empty upload"})
                    return
                if length > _XLSX_MAX_BYTES:
                    self._send_json(413, {
                        "error": f"xlsx grande demais (limite "
                                 f"{_XLSX_MAX_BYTES // (1024*1024)} MB)",
                    })
                    return
                with self.require_session().lock:
                    rdb = self.sess().rdb
                if rdb is None:
                    self._send_json(409, {"error": "RDB nao carregado"})
                    return
                xlsx_bytes = self.rfile.read(length)
                try:
                    updates = parse_xlsx_to_updates(xlsx_bytes)
                except Exception as e:
                    logger.exception("falha parseando xlsx: %s", e)
                    self._send_json(422, {"error": f"xlsx invalido: {e}"})
                    return
                if not updates:
                    self._send_json(422, {
                        "error": "nenhum update valido encontrado no xlsx "
                                 "(verifique A1='Relay:' e A2='GLE:' em cada aba)",
                    })
                    return
                # The source RDB lives in the content cache, which is
                # shared; the derived output belongs to this session.
                out_path = self.sdir("out") / _with_suffix_before_ext(
                    Path(rdb.display_name), "_gle_comments_updated",
                ).name
                try:
                    result = apply_xlsx_updates_to_rdb(
                        rdb_info=rdb, xlsx_updates=updates, output_path=out_path,
                        job=job,
                    )
                except Exception as e:
                    logger.exception("falha aplicando updates: %s", e)
                    self._send_json(500, {"error": str(e)})
                    return
                if not result.get("ok"):
                    # All-or-nothing: no selection was written, so there
                    # is no file to download nor to publish into the library.
                    msg = result.get("error") or "Falha ao aplicar as alterações."
                    job.fail(msg)
                    logger.warning("[gle-exporter] import abortado: %s", msg)
                    self._send_json(400, result)
                    return
                out_abs = Path(result["output_path"]).resolve()
                # The RDB with the comments applied is the next tool's
                # input: it enters the project library, already extracted.
                result["project_file"] = self.publish_output(
                    out_abs, "GLE Exporter", job=job, logger=logger)
                job.finish("RDB atualizado")
                result["download_url"] = self.mount_prefix + "/download?file=" + quote(str(out_abs), safe="")
                result["output_name"] = out_abs.name
                logger.info(
                    "[gle-exporter] import: %d (relay,gle) processado(s), "
                    "%d ok, %d falha -> %s",
                    len(updates), result.get("succeeded", 0),
                    result.get("failed", 0), out_abs.name,
                )
                self._send_json(200, result)
                return

            self._send(404, "not found", "text/plain")

    return Handler
